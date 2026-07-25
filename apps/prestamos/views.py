"""
Autor: Steve
"""
from io import BytesIO
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.catalogo.models import LibroDigital
from apps.core.funcional import (
    contar_prestamos_por_estado,
    libros_mas_prestados,
    total_licencias,
)
from apps.usuarios.decoradores import rol_requerido
from .models import PrestamoDigital


def _obtener_datos_estadisticas():
    """Agrupa en un solo lugar los datos usados por la vista y los reportes."""
    libros = list(LibroDigital.objects.filter(activo=True))
    prestamos = list(PrestamoDigital.objects.all())

    return {
        'libros': libros,
        'prestamos': prestamos,
        'total_licencias': total_licencias(libros),
        'prestamos_por_estado': contar_prestamos_por_estado(prestamos),
        'top_libros': libros_mas_prestados(libros, top=5),
    }


@login_required
def solicitar_prestamo(request, libro_id):
    """
    Cualquier usuario logueado puede solicitar un préstamo, pero la
    solicitud queda siempre pendiente para que el bibliotecario la
    apruebe o rechace desde su panel de gestión.
    """
    libro = get_object_or_404(LibroDigital, pk=libro_id, activo=True)

    # Bloqueo explícito: un usuario suspendido no puede ni siquiera
    # generar la solicitud, más allá de lo que diga el motor de reglas.
    if request.user.suspendido:
        messages.error(request, 'Tu cuenta está suspendida, no podés solicitar préstamos.')
        return redirect('catalogo:lista')

    if request.method == 'POST':
        PrestamoDigital.objects.create(
            libro=libro, usuario=request.user, estado=PrestamoDigital.Estado.PENDIENTE
        )
        messages.warning(
            request,
            f'Solicitud registrada de "{libro.titulo}", pendiente de revisión del bibliotecario.',
        )
        return redirect('catalogo:lista')

    return render(request, 'prestamos/solicitar.html', {'libro': libro})


@login_required
def mis_prestamos(request):
    """Historial completo de préstamos del usuario logueado (todos los estados)."""
    prestamos = request.user.prestamos.all()
    return render(request, 'prestamos/mis_prestamos.html', {'prestamos': prestamos})


@login_required
def devolver_prestamo(request, prestamo_id):
    """
    Permite al propio usuario devolver un libro que tiene aprobado.
    Se filtra por 'usuario=request.user' para que nadie pueda devolver
    préstamos de otra persona manipulando la URL.
    """
    prestamo = get_object_or_404(PrestamoDigital, pk=prestamo_id, usuario=request.user)
    if request.method == 'POST':
        if prestamo.estado == PrestamoDigital.Estado.APROBADO:
            prestamo.devolver()
            messages.success(
                request, f'Devolviste "{prestamo.libro.titulo}". La licencia quedó disponible de nuevo.'
            )
    return redirect('prestamos:mis_prestamos')


@login_required
def leer_libro(request, prestamo_id):
    """
    Muestra el detalle del libro (portada, sinopsis) y el archivo
    embebido para leerlo. Solo se puede acceder si el préstamo es
    propio y está en estado APROBADO: si lo devolviste o nunca te lo
    aprobaron, no deberías poder seguir leyéndolo.
    """
    prestamo = get_object_or_404(
        PrestamoDigital, pk=prestamo_id, usuario=request.user, estado=PrestamoDigital.Estado.APROBADO
    )
    return render(request, 'prestamos/leer.html', {'prestamo': prestamo})


@rol_requerido('BIBLIOTECARIO')
def gestionar_prestamos(request):
    """Panel del bibliotecario con los préstamos pendientes de revisión."""
    pendientes = PrestamoDigital.objects.filter(estado=PrestamoDigital.Estado.PENDIENTE)
    return render(request, 'prestamos/gestionar.html', {'prestamos': pendientes})


@rol_requerido('BIBLIOTECARIO')
def aprobar_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(PrestamoDigital, pk=prestamo_id)
    prestamo.aprobar(request.user)
    messages.success(request, 'Préstamo aprobado.')
    return redirect('prestamos:gestionar')


@rol_requerido('BIBLIOTECARIO')
def rechazar_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(PrestamoDigital, pk=prestamo_id)
    prestamo.rechazar(request.user)
    messages.warning(request, 'Préstamo rechazado.')
    return redirect('prestamos:gestionar')


@rol_requerido('BIBLIOTECARIO')
def estadisticas(request):
    """
    Dashboard con métricas del sistema. Acá es donde se usan las
    funciones puras de core/funcional.py (programación funcional).
    """
    return render(request, 'prestamos/estadisticas.html', _obtener_datos_estadisticas())


@rol_requerido('BIBLIOTECARIO')
def exportar_estadisticas_excel(request):
    """Genera un archivo Excel con el resumen de uso y el ranking de libros."""
    datos = _obtener_datos_estadisticas()

    libro_excel = Workbook()
    hoja_resumen = libro_excel.active
    hoja_resumen.title = 'Resumen'

    # Estilos formales para que el archivo se vea como un reporte institucional.
    color_titulo = '1F4E78'
    color_encabezado = 'D9EAF7'
    borde_fino = Border(
        left=Side(style='thin', color='B7C9D6'),
        right=Side(style='thin', color='B7C9D6'),
        top=Side(style='thin', color='B7C9D6'),
        bottom=Side(style='thin', color='B7C9D6'),
    )

    hoja_resumen.column_dimensions['A'].width = 28
    hoja_resumen.column_dimensions['B'].width = 14
    hoja_resumen.column_dimensions['C'].width = 14
    hoja_resumen['A1'] = 'Reporte de uso'
    hoja_resumen['A1'].font = Font(name='Times New Roman', size=16, bold=True, color=color_titulo)
    hoja_resumen['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    hoja_resumen['A2'].font = Font(name='Times New Roman', size=10, italic=True, color='666666')

    hoja_resumen['A3'] = 'Licencias totales'
    hoja_resumen['B3'] = datos['total_licencias']
    hoja_resumen['A3'].font = Font(name='Times New Roman', bold=True)
    hoja_resumen['B3'].font = Font(name='Times New Roman')
    hoja_resumen['A3'].fill = PatternFill('solid', fgColor=color_encabezado)
    hoja_resumen['B3'].fill = PatternFill('solid', fgColor='F7FBFF')
    hoja_resumen['A3'].border = borde_fino
    hoja_resumen['B3'].border = borde_fino

    fila = 5
    hoja_resumen[f'A{fila}'] = 'Préstamos por estado'
    hoja_resumen[f'A{fila}'].font = Font(name='Times New Roman', bold=True, size=12, color=color_titulo)
    fila += 1
    for estado, cantidad in datos['prestamos_por_estado'].items():
        hoja_resumen[f'A{fila}'] = estado
        hoja_resumen[f'B{fila}'] = cantidad
        hoja_resumen[f'A{fila}'].font = Font(name='Times New Roman')
        hoja_resumen[f'B{fila}'].font = Font(name='Times New Roman')
        hoja_resumen[f'A{fila}'].border = borde_fino
        hoja_resumen[f'B{fila}'].border = borde_fino
        hoja_resumen[f'A{fila}'].alignment = Alignment(horizontal='left')
        hoja_resumen[f'B{fila}'].alignment = Alignment(horizontal='center')
        fila += 1

    hoja_top = libro_excel.create_sheet('Top libros')
    hoja_top.column_dimensions['A'].width = 10
    hoja_top.column_dimensions['B'].width = 42
    hoja_top.column_dimensions['C'].width = 28
    hoja_top['A1'] = 'Top 5 libros más prestados'
    hoja_top['A1'].font = Font(name='Times New Roman', size=16, bold=True, color=color_titulo)
    hoja_top['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    hoja_top['A2'].font = Font(name='Times New Roman', size=10, italic=True, color='666666')
    hoja_top['A3'] = 'Puesto'
    hoja_top['B3'] = 'Libro'
    hoja_top['C3'] = 'Autor'
    for celda in ('A3', 'B3', 'C3'):
        hoja_top[celda].font = Font(name='Times New Roman', bold=True, color='1F1F1F')
        hoja_top[celda].fill = PatternFill('solid', fgColor=color_encabezado)
        hoja_top[celda].border = borde_fino
        hoja_top[celda].alignment = Alignment(horizontal='center', vertical='center')

    for indice, libro in enumerate(datos['top_libros'], start=1):
        fila_libro = indice + 3
        hoja_top[f'A{fila_libro}'] = indice
        hoja_top[f'B{fila_libro}'] = libro.titulo
        hoja_top[f'C{fila_libro}'] = libro.autor
        for celda in (f'A{fila_libro}', f'B{fila_libro}', f'C{fila_libro}'):
            hoja_top[celda].font = Font(name='Times New Roman')
            hoja_top[celda].border = borde_fino
            hoja_top[celda].alignment = Alignment(vertical='center')
        hoja_top[f'A{fila_libro}'].alignment = Alignment(horizontal='center', vertical='center')

    respuesta = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    respuesta['Content-Disposition'] = 'attachment; filename="reporte_estadisticas.xlsx"'
    libro_excel.save(respuesta)
    return respuesta


@rol_requerido('BIBLIOTECARIO')
def exportar_estadisticas_pdf(request):
    """Genera un PDF breve con las mismas métricas mostradas en pantalla."""
    datos = _obtener_datos_estadisticas()
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        'TituloReporte',
        parent=estilos['Title'],
        fontName='Times-Bold',
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=8,
    )
    subtitulo = ParagraphStyle(
        'SubtituloReporte',
        parent=estilos['Normal'],
        fontName='Times-Italic',
        fontSize=10,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#666666'),
        spaceAfter=14,
    )
    seccion = ParagraphStyle(
        'SeccionReporte',
        parent=estilos['Heading2'],
        fontName='Times-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1F1F1F'),
        spaceBefore=8,
        spaceAfter=6,
    )
    normal = ParagraphStyle(
        'NormalReporte',
        parent=estilos['Normal'],
        fontName='Times-Roman',
        fontSize=10.5,
        leading=13,
        alignment=TA_LEFT,
        textColor=colors.black,
    )

    tabla_estadisticas = [
        [Paragraph('Concepto', seccion), Paragraph('Valor', seccion)],
        [Paragraph('Licencias totales', normal), Paragraph(str(datos['total_licencias']), normal)],
    ]
    for estado, cantidad in datos['prestamos_por_estado'].items():
        tabla_estadisticas.append([Paragraph(estado, normal), Paragraph(str(cantidad), normal)])

    tabla_top = [[Paragraph('Puesto', seccion), Paragraph('Libro', seccion), Paragraph('Autor', seccion)]]
    for indice, libro in enumerate(datos['top_libros'], start=1):
        tabla_top.append([
            Paragraph(str(indice), normal),
            Paragraph(libro.titulo, normal),
            Paragraph(libro.autor, normal),
        ])

    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9EAF7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F1F1F')),
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEADING', (0, 0), (-1, -1), 13),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor('#B7C9D6')),
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#8EA9C1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FBFF')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])

    resumen = Table(tabla_estadisticas, colWidths=[110 * mm, 40 * mm], repeatRows=1)
    resumen.setStyle(estilo_tabla)

    ranking = Table(tabla_top, colWidths=[20 * mm, 90 * mm, 70 * mm], repeatRows=1)
    ranking.setStyle(estilo_tabla)

    def encabezado_pie(canvas_obj, doc_obj):
        # Encabezado y pie fijos para dar aspecto institucional al PDF.
        canvas_obj.saveState()
        canvas_obj.setStrokeColor(colors.HexColor('#1F4E78'))
        canvas_obj.setLineWidth(1)
        canvas_obj.line(doc_obj.leftMargin, doc_obj.height + doc_obj.topMargin + 6, doc_obj.pagesize[0] - doc_obj.rightMargin, doc_obj.height + doc_obj.topMargin + 6)
        canvas_obj.setFont('Times-Bold', 10)
        canvas_obj.drawRightString(doc_obj.pagesize[0] - doc_obj.rightMargin, 10 * mm, f'Página {doc_obj.page}')
        canvas_obj.setFont('Times-Roman', 9)
        canvas_obj.drawString(doc_obj.leftMargin, 10 * mm, 'Biblioteca Virtual')
        canvas_obj.restoreState()

    historia = [
        Paragraph('Reporte de uso de Biblioteca Virtual', titulo),
        Paragraph(f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}', subtitulo),
        Paragraph('Resumen general', seccion),
        resumen,
        Spacer(1, 10),
        Paragraph('Top 5 libros más prestados', seccion),
        ranking,
    ]

    doc.build(historia, onFirstPage=encabezado_pie, onLaterPages=encabezado_pie)

    respuesta = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    respuesta['Content-Disposition'] = 'attachment; filename="reporte_estadisticas.pdf"'
    return respuesta